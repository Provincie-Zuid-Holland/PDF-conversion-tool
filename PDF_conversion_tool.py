#!/usr/bin/env python
# coding: utf-8
# Author: Joana Cardoso

import json
import logging
import os
import tkinter as tk
from openpyxl import Workbook
from pathlib import Path
from tkinter import messagebox, filedialog, ttk
import _functions.unzip_files as uz
import _functions.check_length as cl
import _functions.get_files as gf
import _functions.duplicates as dup
import _functions.print as prt
import _functions.combine as comb

logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s %(levelname)s %(funcName)s %(message)s',
                    filename='Logging_PDF_conversion_tool.log'
                    )  # to see log in console remove filename


class Application(tk.Frame):
    """
    This class is a python gui to convert files to pdf
    Provides functionality such as:
    - Selecting de file location (including unzipping of zip files)
    - Determining the amount of files available for conversion
    - Extracting and converting emails and attachments to pdf
    - Combining emails and attachments into 1 pdf
    - Converting individual files to pdf
    - Removing doubles
    - Unlocking blocked pdf files
    - Making a log file where all the steps can be followed and errors can be traced

    Methods
    -------
    folder_type()
        Sets up a screen to select the location of the pdf documents for unlocking
    select_zip()
        Calls _functions.unzip_file.unzip_files and gets the path to the selected file
    select_folder()
        Gets the path to the selected folder and calls _functions.check_length.check_length
    find_files()
        Gets the amount of files found in the selected folder or zip file and presents the results in a screen
    duplicates_yn()
        Opens a screen to select whether duplicates should be removed or not
    open_folder(path: str)
        Opens the links in the last screen of the tool
    convert_files(remove_duplicates: bool)
        Processes the selected files and calls several other functions
    cancel()
        Destroys the tool if the tool is halfway stopped
    ready()
        Destroys the tool if the tool is ready with the conversion
    """
    logging.info('Starting Tool')

    def __init__(self, parent):
        """
        Initializes the tool and sets up a start screen.
        """
        root.withdraw()
        self.progress = None
        self.style = None
        tk.Frame.__init__(self, master=parent)
        info = 'This application converts files to PDF.\n\nSelect a folder to start converting.\n\nThis application is developed by the Data Wharehouse team of the Province of Zuid-Holland, The Netherlands.'
        self.parent = parent
        parent.iconbitmap('logo.ico')
        messagebox.showinfo(title=None, message=info)
        self.folder_type()

    def folder_type(self):
        """
        Sets up a screen to select the location of the pdf documents for unlocking
        """
        self.parent = tk.Tk()
        w = self.parent.winfo_reqwidth()
        h = self.parent.winfo_reqheight()
        ws = self.parent.winfo_screenwidth()
        hs = self.parent.winfo_screenheight()
        x = (ws / 2.3) - (w / 2.3)
        y = (hs / 2) - (h / 2)
        self.parent.geometry('+%d+%d' % (x, y))
        self.parent.iconbitmap('logo.ico')
        self.parent.title('File type')
        text = 'What do you want to select?'
        self.parent.resizable(width="false", height="false")
        self.parent.minsize(width=250, height=75)
        self.parent.maxsize(width=250, height=75)
        self.label = tk.Label(self.parent, text=text).place(
            relx=.1, rely=.2, anchor="w")
        self.button1 = tk.Button(self.parent, text='Folder', command=self.select_folder).place(
            relx=.38, rely=.7, anchor="c")
        self.button2 = tk.Button(self.parent, text='Zipped folder',
                                 command=self.select_zip).place(relx=.64, rely=.7, anchor="c")
        self.quit = tk.Button(self.parent, text='Stop', command=self.cancel).place(
            relx=.88, rely=.7, anchor="c")
        self.parent.protocol("WM_DELETE_WINDOW", self.cancel)

    def select_zip(self):
        """
        If the option zip file is selected in the folder_type screen, this method calls the file _functions.unzip_file.unzip_files and 
            gets the path to the selected file
        """
        self.parent.destroy()
        zip_dir = filedialog.askopenfilename(initialdir="/Users", title="Zipped folder selection",
                                             filetypes=(("ZIP files", "*.ZIP"), ("zip files", "*.zip")))
        if not zip_dir:
            logging.info('Back to selection folder type')
            self.folder_type()

        else:
            self.zip_dir = os.path.abspath(zip_dir)
            logging.info(f'Zip_dir: {zip_dir}')
            self.process_dir = os.path.abspath(os.path.splitext(zip_dir)[0])
            logging.info(f'Process_dir: {self.process_dir}')
            logging.info(f'Started unzipping folder: {zip_dir}')
            uz.unzip_files(self.zip_dir, self.process_dir)
            self.find_files()

    def select_folder(self):
        """
        If the option folder is selected in the folder_type screen, this method gets the path to the selected folder
        It calls the function check_length in file _functions.check_length
        If necessary calls the function unzip_files in file _functions.unzip_file
        """
        self.parent.destroy()
        self.process_dir = filedialog.askdirectory(
            initialdir="/Users", title="Map selectie")

        if not self.process_dir:
            logging.info('Back to select folder type')
            self.folder_type()

        else:
            self.process_dir = os.path.abspath(self.process_dir)
            logging.info(f'Process directory: {self.process_dir}')
            for root, dirs, files in os.walk(self.process_dir):
                for file in files:
                    des_dir = os.path.join(root, os.path.dirname(file))
                    file_path = os.path.join(root, file)
                    logging.info(f'File: {file}')
                    logging.info(f'Destination directory: {des_dir}')
                    file_name, long_name = cl.check_length(
                        des_dir=des_dir, file=file)

                    if long_name == True:
                        logging.info(f'Long name')
                        try:
                            os.rename(file_path, file_name)
                            logging.debug(
                                f'Changing name: {file} into {file_name}')
                        except:
                            logging.error(
                                f'Failed to change name: {file} into {file_name}')
                    name_lower = file.lower()
                    if name_lower.endswith('.zip'):
                        zip_dir = os.path.join(root, file)
                        proc_zip = os.path.abspath(
                            os.path.splitext(zip_dir)[0])
                        logging.info(f'Zip_dir: {zip_dir}')
                        logging.info(f'Proc dir: {proc_zip}')
                        uz.unzip_files(zip_dir, proc_zip)
                        try:
                            os.remove(zip_dir)
                            logging.debug(f'Removed zip: {zip_dir}')
                        except:
                            logging.error(f'Failed to remove zip: {zip_dir}')
            self.find_files()

    def find_files(self):
        """
        Gets the amount of files found in the selected folder or zip file and presents the results on a screen
        """
        logging.info('Find files')

        # count files (emails + other files)
        email_files = []
        other_files = 0
        self.pdf_files = 0
        self.all_files = []
        self.empty_dir = []

        for root, dirs, files in os.walk(self.process_dir):
            if not len(dirs) and not len(files):
                # Adding the empty directory to list
                self.empty_dir.append(root)
            for name in files:
                name_lower = name.lower()
                if not name_lower.endswith(".zip"):
                    if name_lower.endswith((".msg", ".eml")):
                        email_files.append(os.path.join(root, name))
                    elif name_lower.endswith(".pdf"):
                        self.pdf_files += 1
                    else:
                        other_files += 1
                    self.all_files.append(os.path.join(root, name))

        # print messages
        total_found = 'Number of files found: ' + str(len(self.all_files))
        found_emails = "Number of emails: " + \
            str(len(email_files))
        found_pdfs = 'Number of pdf\'s: ' + str(self.pdf_files)
        found_others = "Other files: " + str(other_files)

        self.parent = tk.Tk()
        w = self.parent.winfo_reqwidth()
        h = self.parent.winfo_reqheight()
        ws = self.parent.winfo_screenwidth()
        hs = self.parent.winfo_screenheight()
        x = (ws / 2.3) - (w / 2.3)
        y = (hs / 2) - (h / 2)
        self.parent.geometry('+%d+%d' % (x, y))
        self.parent.iconbitmap('logo.ico')
        self.parent.title('Found files')
        self.parent.resizable(width="false", height="false")
        self.parent.minsize(width=300, height=150)
        self.parent.maxsize(width=300, height=150)
        self.label = tk.Label(self.parent, text=total_found).place(
            relx=.1, rely=.1, anchor="w")
        self.label = tk.Label(self.parent, text=found_emails).place(
            relx=.1, rely=.3, anchor="w")
        self.label = tk.Label(self.parent, text=found_pdfs).place(
            relx=.1, rely=.5, anchor="w")
        self.label = tk.Label(self.parent, text=found_others).place(
            relx=.1, rely=.7, anchor="w")

        self.contin = tk.Button(self.parent, text='Continue', command=self.duplicates_yn).place(
            relx=.72, rely=.9, anchor="c")
        self.quit = tk.Button(self.parent, text='Stop', command=self.cancel).place(
            relx=.88, rely=.9, anchor="c")
        self.parent.protocol("WM_DELETE_WINDOW", self.cancel)

    def duplicates_yn(self):
        """
        Opens a screen to select whether duplicates should be removed or not
        """
        self.parent.destroy()

        self.parent = tk.Tk()
        w = self.parent.winfo_reqwidth()
        h = self.parent.winfo_reqheight()
        ws = self.parent.winfo_screenwidth()
        hs = self.parent.winfo_screenheight()
        x = (ws / 2.3) - (w / 2.3)
        y = (hs / 2) - (h / 2)
        self.parent.geometry('+%d+%d' % (x, y))
        self.parent.iconbitmap('logo.ico')
        self.parent.title('Remove duplicates')
        text = 'Would you like to remove duplicates?'
        self.parent.resizable(width="false", height="false")
        self.parent.minsize(width=250, height=75)
        self.parent.maxsize(width=250, height=75)
        self.label = tk.Label(self.parent, text=text).place(
            relx=.1, rely=.2, anchor="w")
        self.remov_yes = tk.Button(self.parent, text='Yes', command=lambda: self.convert_files(
            True)).place(relx=.55, rely=.7, anchor="c")
        self.remov_no = tk.Button(self.parent, text='No', command=lambda: self.convert_files(
            False)).place(relx=.67, rely=.7, anchor="c")
        self.quit = tk.Button(self.parent, text='Stop', command=self.cancel).place(
            relx=.81, rely=.7, anchor="c")
        self.parent.protocol("WM_DELETE_WINDOW", self.cancel)

    def open_folder(self, path):
        """
        Opens the links in the last screen of the tool

        Parameters
        ----------
        path: str
            The path to the link
        """
        os.startfile(path, 'open')

    def convert_files(self, remove_duplicates):
        """
        Processes the selected files and calls several other methods.

        Parameters
        ----------
        remove_duplicates: bool
            Remove_duplicates is set to True or False, depending whether duplicates are to be removed or not
        """
        self.parent.destroy()

        # create progress bar
        self.parent = tk.Tk()
        w = self.parent.winfo_reqwidth()
        h = self.parent.winfo_reqheight()
        ws = self.parent.winfo_screenwidth()
        hs = self.parent.winfo_screenheight()
        x = (ws / 2) - (w / 2)
        y = (hs / 2) - (h / 2)
        self.parent.geometry('+%d+%d' % (x, y))
        self.parent.iconbitmap('logo.ico')
        self.parent.title('')

        self.style = ttk.Style(self.parent)
        self.style.layout('text.Horizontal.TProgressbar',
                          [('Horizontal.Progressbar.trough',
                            {'children': [('Horizontal.Progressbar.pbar',
                                           {'side': 'left', 'sticky': 'ns'})],
                             'sticky': 'nswe'}),
                           ('Horizontal.Progressbar.label', {'sticky': ''})])

        self.progress = ttk.Progressbar(
            self.parent,
            orient=tk.HORIZONTAL,
            mode="determinate",
            length=300,
            maximum=100,
            style='text.Horizontal.TProgressbar'
        )
        self.progress.grid(row=1, column=1, columnspan=2, padx=5, pady=5)
        self.label = ttk.Label(
            self.parent, text='Getting files and extracting emails')
        self.label.grid(row=2, column=1, pady=5, padx=5, sticky='nswe')
        self.button = ttk.Button(
            self.parent, text='Stop', command=self.cancel)
        self.button.grid(row=3, column=2, pady=5, padx=5, sticky='e')
        self.parent.protocol("WM_DELETE_WINDOW", self.cancel)

        # set output directory
        out_dir = os.path.join(os.path.dirname(
            self.process_dir), "PDF_" + os.path.basename(self.process_dir))
        logging.info(f'Output directory: {out_dir}')

        # set start value Bestanden ophalen en emails extraheren
        logging.info(f'Set start value Bestanden ophalen en emails extraheren')
        self.progress['value'] = 1
        self.style.configure('text.Horizontal.TProgressbar',
                             text='{:g} %'.format(self.progress['value']))

        # print email body to pdf, extract attachments and copy other files to output folder
        logging.info(
            'Started converting files to PDF and extract email attachments')
        msg_list, file_log = gf.get_all_files(
            files=self.all_files, process_dir=self.process_dir, out_dir=out_dir, progress=self.progress, style=self.style)
        logging.info(
            'Finished converting files to PDF and extract email attachments')

        # update status pb Bestanden ophalen en emails extraheren
        logging.info(
            f'Update status pb Bestanden ophalen en emails extraheren')
        self.progress.update()
        self.progress.update_idletasks()

        logging.info(f'Remove duplicates is set to {remove_duplicates}')
        if remove_duplicates == True:
            # set start value Dubbelingen verwijderen
            self.progress['value'] = 1
            self.label.configure(text="Removing duplicates")
            self.style.configure('text.Horizontal.TProgressbar',
                                 text='{:g} %'.format(self.progress['value']))

            # remove duplicates
            logging.info('Started removing duplicate files')
            out_duplicates, duplicates, pdf_duplos = dup.remove_duplicates(
                process_dir=self.process_dir, out_dir=out_dir, progress=self.progress, style=self.style, file_log=file_log, remove_duplicates=remove_duplicates)
            logging.info('Finished removing duplicate files')

            # update status pb Dubbelingen verwijderen
            self.progress.update()
            self.progress.update_idletasks()

        else:
            # look for duplicates but do not move
            logging.info('Started looking for duplicate files')
            out_duplicates, duplicates, pdf_duplos = dup.remove_duplicates(
                process_dir=self.process_dir, out_dir=out_dir, progress=self.progress, style=self.style, file_log=file_log, remove_duplicates=remove_duplicates)
            logging.info('Finished looking for duplicate files')

        # set start value Bestanden printen naar PDF
        self.progress['value'] = 1
        self.label.configure(text="Printing files to pdf")
        self.style.configure('text.Horizontal.TProgressbar',
                             text='{:g} %'.format(self.progress['value']))

        # print separate files and attachments
        logging.info('Started printing files to PDF')
        prt.print_to_pdf(out_dir=out_dir, file_log=file_log,
                         progress=self.progress, style=self.style)
        logging.info('Finished printing files to PDF')

        # update status pb Bestanden printen naar PDF
        self.progress.update()
        self.progress.update_idletasks()

        # set start value PDF's combineren
        self.progress['value'] = 1
        self.label.configure(text="Combining pdf's")
        self.style.configure('text.Horizontal.TProgressbar',
                             text='{:g} %'.format(self.progress['value']))

        # combine pdf's
        logging.info('Started combining PDF files')
        comb.combine_pdf(process_dir=self.process_dir, out_dir=out_dir,
                         msg_list=msg_list, progress=self.progress, style=self.style)
        logging.info('Finished combining PDF files')

        # update status pb PDF's combineren
        self.progress.update()
        self.progress.update_idletasks()

        # make empty directories if they exist in the original directory
        if len(self.empty_dir) > 0:
            logging.info('Started creating empty directories')
            for emp_dir in self.empty_dir:
                create_emp_dir = emp_dir.replace(self.process_dir, out_dir)
                try:
                    # Create empty directory in output dir
                    Path(create_emp_dir).mkdir(parents=True, exist_ok=True)
                    logging.debug(
                        f'Creating empty directory: {create_emp_dir}')
                except:
                    logging.error(
                        f'Failed to create empty directory: {create_emp_dir}')
            logging.info('Finished creating empty directories')

        # count converted emails
        total_pdf = 0
        not_converted = []
        excel_html = []
        ext = ['.xlsx', '.xls', '.pdf', ".htm", ".html", ".XLS", ".XLXS"]
        for root, dirs, files in os.walk(out_dir):
            for name in files:
                name_lower = name.lower()
                file_path = os.path.join(root, name_lower)
                if name_lower.endswith('.pdf'):
                    total_pdf += 1
                if name_lower.endswith(".xlsx") or name_lower.endswith(".xls") or name_lower.endswith(".htm") or name_lower.endswith(".html"):
                    excel_html.append(name_lower)
                if not name_lower.endswith(tuple(ext)):
                    not_converted.append(name_lower)
                    logging.info(f'Not converted file: {file_path}')
                    try:
                        d = next(
                            item for item in file_log if item['File name'] == file_path)
                        d['Check'] = True
                        d['Combined'] = False
                        logging.debug(
                            f'Updating log table for not converted file: {file_path}')
                    except:
                        logging.debug(
                            f'Failed to update log table for not converted file: {file_path}')

        total_pdf = total_pdf + pdf_duplos

        # Make log table json (for debugging)
        try:
            with open('Log_results.json', 'w') as f:
                json.dump(file_log, f)
                logging.debug(f'Making json table Log_results')
        except:
            logging.debug(f'Failed to make json table Log_results')

        # Make log table excel (for end-user)
        Log_table_name = 'Results_PDF_' + \
            str(os.path.basename(self.process_dir)) + '.xlsx'
        Log_table_path = os.path.join(
            os.path.dirname(self.process_dir), Log_table_name)
        # delete 2 columns
        logging.info('Deleting columns and replacing values')
        for d in file_log:
            try:
                del (d['File name'])
                del (d['File'])
                d.update((k, "Y") for k, v in d.items() if v == True)
                d.update((k, "N") for k, v in d.items() if v == False)
            except:
                logging.error('Failed to delete columns and replace values')

        try:
            wb = Workbook()
            wb.save(Log_table_path)
            wb = Workbook(write_only=True)
            sheet = wb.create_sheet('Results')
            fieldnames = list(file_log[0])
            sheet.append(fieldnames)

            for x in file_log:
                sheet.append(list(x.values()))
            wb.save(Log_table_path)
        except:
            logging.error(f'Failed to make log table: {Log_table_path}')

        # print messages
        klaar = 'Ready with conversion to pdf'
        total_converted = 'Total converted files: ' + \
            str(total_pdf - self.pdf_files)
        total_not_convert = 'Not converted files: ' + \
            str(len(not_converted))
        output = 'Converted files are in folder: ' + out_dir
        output2 = 'Results are in file: ' + Log_table_path
        output3 = 'Duplicates are in folder: ' + out_duplicates
        output4 = 'Check if all files were converted.'

        if remove_duplicates == True:
            if len(duplicates) > 0:
                self.parent.destroy()
                self.parent = tk.Tk()
                w = self.parent.winfo_reqwidth()
                h = self.parent.winfo_reqheight()
                ws = self.parent.winfo_screenwidth()
                hs = self.parent.winfo_screenheight()
                x = (ws / 3) - (w / 3)
                y = (hs / 2.3) - (h / 2.3)
                self.parent.geometry('+%d+%d' % (x, y))
                self.parent.iconbitmap('logo.ico')
                self.parent.title('Converted files')
                self.parent.resizable(width="false", height="false")
                self.parent.minsize(width=800, height=250)
                self.parent.maxsize(width=800, height=250)
                self.label = tk.Label(self.parent, text=klaar).place(
                    relx=.03, rely=.1, anchor="w")
                self.label = tk.Label(self.parent, text=total_converted).place(
                    relx=.03, rely=.22, anchor="w")
                self.label = tk.Label(self.parent, text=total_not_convert).place(
                    relx=.03, rely=.34, anchor="w")

                # Define clickable labels
                label1 = tk.Label(self.parent, text=output,
                                  fg='blue', cursor='hand2')
                label1.pack()
                label1.bind("<Button-1>", lambda e: self.open_folder(out_dir))
                self.label1 = label1.place(relx=.03, rely=.49, anchor="w")

                label2 = tk.Label(self.parent, text=output3,
                                  fg='blue', cursor='hand2')
                label2.pack()
                label2.bind(
                    "<Button-1>", lambda e: self.open_folder(out_duplicates))
                self.label2 = label2.place(relx=.03, rely=.59, anchor="w")

                label3 = tk.Label(self.parent, text=output2,
                                  fg='blue', cursor='hand2')
                label3.pack()
                label3.bind(
                    "<Button-1>", lambda e: self.open_folder(Log_table_path))
                self.label3 = label3.place(relx=.03, rely=.69, anchor="w")

                self.label = tk.Label(self.parent, text=output4).place(
                    relx=.03, rely=.79, anchor="w")
                self.quit = tk.Button(self.parent, text='Close', command=self.ready).place(
                    relx=.95, rely=.9, anchor="c")
                self.parent.protocol("WM_DELETE_WINDOW", self.ready)

            else:
                self.parent.destroy()
                self.parent = tk.Tk()
                w = self.parent.winfo_reqwidth()
                h = self.parent.winfo_reqheight()
                ws = self.parent.winfo_screenwidth()
                hs = self.parent.winfo_screenheight()
                x = (ws / 3) - (w / 3)
                y = (hs / 2.3) - (h / 2.3)
                self.parent.geometry('+%d+%d' % (x, y))
                self.parent.iconbitmap('logo.ico')
                self.parent.title('Converted files')
                self.parent.resizable(width="false", height="false")
                self.parent.minsize(width=800, height=250)
                self.parent.maxsize(width=800, height=250)
                self.label = tk.Label(self.parent, text=klaar).place(
                    relx=.1, rely=.1, anchor="w")
                self.label = tk.Label(self.parent, text=total_converted).place(
                    relx=.1, rely=.22, anchor="w")
                self.label = tk.Label(self.parent, text=total_not_convert).place(
                    relx=.1, rely=.34, anchor="w")

                # Define clickable labels
                label1 = tk.Label(self.parent, text=output,
                                  fg='blue', cursor='hand2')
                label1.pack()
                label1.bind("<Button-1>", lambda e: self.open_folder(out_dir))
                self.label1 = label1.place(relx=.1, rely=.49, anchor="w")

                label2 = tk.Label(self.parent, text=output2,
                                  fg='blue', cursor='hand2')
                label2.pack()
                label2.bind(
                    "<Button-1>", lambda e: self.open_folder(Log_table_path))
                self.label2 = label2.place(relx=.1, rely=.59, anchor="w")

                self.label = tk.Label(self.parent, text=output4).place(
                    relx=.1, rely=.69, anchor="w")
                self.quit = tk.Button(self.parent, text='Close', command=self.ready).place(
                    relx=.88, rely=.9, anchor="c")
                self.parent.protocol("WM_DELETE_WINDOW", self.ready)

        else:
            self.parent.destroy()
            self.parent = tk.Tk()
            w = self.parent.winfo_reqwidth()
            h = self.parent.winfo_reqheight()
            ws = self.parent.winfo_screenwidth()
            hs = self.parent.winfo_screenheight()
            x = (ws / 3) - (w / 3)
            y = (hs / 2.3) - (h / 2.3)
            self.parent.geometry('+%d+%d' % (x, y))
            self.parent.iconbitmap('logo.ico')
            self.parent.title('Converted files')
            self.parent.resizable(width="false", height="false")
            self.parent.minsize(width=800, height=250)
            self.parent.maxsize(width=800, height=250)
            self.label = tk.Label(self.parent, text=klaar).place(
                relx=.1, rely=.1, anchor="w")
            self.label = tk.Label(self.parent, text=total_converted).place(
                relx=.1, rely=.22, anchor="w")
            self.label = tk.Label(self.parent, text=total_not_convert).place(
                relx=.1, rely=.34, anchor="w")

            # Define clickable labels
            label1 = tk.Label(self.parent, text=output,
                              fg='blue', cursor='hand2')
            label1.pack()
            label1.bind("<Button-1>", lambda e: self.open_folder(out_dir))
            self.label1 = label1.place(relx=.1, rely=.49, anchor="w")

            label2 = tk.Label(self.parent, text=output2,
                              fg='blue', cursor='hand2')
            label2.pack()
            label2.bind(
                "<Button-1>", lambda e: self.open_folder(Log_table_path))
            self.label2 = label2.place(relx=.1, rely=.59, anchor="w")

            self.label = tk.Label(self.parent, text=output4).place(
                relx=.1, rely=.69, anchor="w")
            self.quit = tk.Button(self.parent, text='Close', command=self.ready).place(
                relx=.88, rely=.9, anchor="c")
            self.parent.protocol("WM_DELETE_WINDOW", self.ready)

    def cancel(self):
        """
        Destroys the tool and inserts info in the log file if the tool is half-way stopped.
        """
        self.parent.destroy()
        self.parent.quit()
        logging.shutdown()
        logging.info('Tool cancelled')

    def ready(self):
        """
        Destroys the tool if the tool is ready with the conversion
        """
        self.parent.destroy()
        self.parent.quit()
        logging.shutdown()
        logging.info('Ready with tool')


if __name__ == '__main__':
    root = tk.Tk()
    app = Application(root)
    root.mainloop()
